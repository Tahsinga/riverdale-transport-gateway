from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from decimal import Decimal
import json
import datetime

from .models import RFIDTag, Student, Account, RideLog, SystemConfig, UnregisteredTag
from django.contrib.auth.models import User
from django.utils import timezone
from django.db.models import Sum
from django.views.decorators.http import require_http_methods
from django.core.cache import cache
from django.conf import settings
import openpyxl
import re
from io import BytesIO
from pathlib import Path


def logo_png(request):
	"""Serve the project logo as a fallback when static files are not yet collected/restarted."""
	paths = [
		Path(settings.BASE_DIR) / 'static' / 'img' / 'logo.png',
		Path(settings.BASE_DIR) / 'staticfiles' / 'img' / 'logo.png',
		Path(settings.BASE_DIR) / 'config' / 'templates' / 'config' / 'logo.png',
	]
	for path in paths:
		if path.exists():
			return FileResponse(open(path, 'rb'), content_type='image/png')
	return JsonResponse({'error': 'logo not found'}, status=404)


def login_view(request):
	"""Handle user login with Django authentication"""
	if request.user.is_authenticated:
		return redirect('config:dashboard')
	
	if request.method == 'POST':
		username = request.POST.get('username', '').strip()
		password = request.POST.get('password', '')
		
		# Authenticate user against Django's auth system
		user = authenticate(request, username=username, password=password)
		
		if user is not None:
			login(request, user)
			messages.success(request, f'Welcome back, {user.username}!')
			return redirect('config:dashboard')
		else:
			messages.error(request, 'Invalid username or password.')
			return render(request, 'config/riverdale_login.html', {
				'username': username,
				'error': 'Invalid credentials'
			})
	
	return render(request, 'config/riverdale_login.html')


def logout_view(request):
	"""Handle user logout"""
	logout(request)
	messages.success(request, 'You have been logged out.')
	return redirect('config:login')


@login_required(login_url='config:login')
def dashboard(request):
	total_rides = RideLog.objects.filter(success=True).count()
	total_revenue = RideLog.objects.filter(success=True).aggregate(sum=Sum('fare'))['sum'] or 0
	recent_rides = RideLog.objects.select_related('student').order_by('-timestamp')[:10]
	students = Student.objects.count()

	chart_data = list(RideLog.objects.filter(success=True).order_by('-timestamp')[:20].values_list('fare', flat=True))

	# build students data for client-side UI
	students_qs = Student.objects.select_related('rfid_tag').all()
	students_list = []
	for s in students_qs:
		acct = Account.objects.filter(student=s).first()
		students_list.append({
			'id': s.id,
			'name': s.name,
			'rfidUid': s.rfid_tag.uid if s.rfid_tag else '',
			'balance': float(acct.balance) if acct else 0,
			'active': bool(s.rfid_tag and s.rfid_tag.assigned),
			# include related user info if present so frontend can show username/email
			'username': s.user.username if getattr(s, 'user', None) else None,
			'email': s.user.email if getattr(s, 'user', None) else None,
			# metadata fields persisted on Student
			'grade': s.grade,
			'gender': s.gender,
			'roll': s.roll,
			'register_no': s.roll,
			'parent': s.parent_contact,
			'section': s.parent_contact,
		})

	recent_qs = RideLog.objects.select_related('student').order_by('-timestamp')[:20]
	recent_list = []
	for r in recent_qs:
		recent_list.append({
			'studentName': r.student.name,
			'amount': float(-r.fare) if r.success else float(-r.fare),
			'description': 'RFID tap',
			# ISO-like timestamp (YYYY-MM-DD HH:MM:SS) for reliable client-side parsing
			'timestamp': timezone.localtime(r.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
			'type': 'payment' if r.success else 'failed'
		})

	return render(request, 'config/dashboard.html', {
		'total_rides': total_rides,
		'total_revenue': total_revenue,
		'recent_rides': recent_rides,
		'students': students,
		'chart_data': chart_data,
		'students_json': json.dumps(students_list),
		'transactions_json': json.dumps(recent_list),
		'cost_per_ride': float(SystemConfig.get_solo().cost_per_ride),
		'min_balance': float(SystemConfig.get_solo().min_balance),
	})


@csrf_exempt
def rfid_scan(request):
	if request.method != 'POST':
		return JsonResponse({'error': 'POST required'}, status=400)
	try:
		payload = json.loads(request.body.decode('utf-8'))
	except Exception:
		payload = request.POST

	uid = payload.get('uid') or payload.get('tag')
	# use configured cost per ride by default
	default_fare = SystemConfig.get_solo().cost_per_ride
	if 'fare' in payload and payload.get('fare') not in (None, '', []):
		try:
			fare = Decimal(str(payload.get('fare')))
		except Exception:
			fare = default_fare
	else:
		fare = default_fare

	if not uid:
		return JsonResponse({'error': 'missing uid'}, status=400)

	tag = RFIDTag.objects.filter(uid=uid).first()
	if not tag:
		# unknown tag - save to database and cache for dashboard polling
		UnregisteredTag.objects.create(uid=uid)
		cache.set('last_scan', {
			'uid': uid,
			'student': None,
			'status': 'tag_not_found',
			'remaining': None,
			'timestamp': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
		}, 30)
		return JsonResponse({'error': 'tag not found'}, status=404)
	if not tag.assigned:
		# tag exists but not assigned - save to database and cache
		UnregisteredTag.objects.create(uid=uid)
		cache.set('last_scan', {
			'uid': uid,
			'student': None,
			'status': 'tag_unassigned',
			'remaining': None,
			'timestamp': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
		}, 30)
		return JsonResponse({'error': 'tag unassigned'}, status=404)

	student = Student.objects.filter(rfid_tag=tag).first()
	if not student:
		# tag assigned but student not found - save to database and cache
		UnregisteredTag.objects.create(uid=uid)
		cache.set('last_scan', {
			'uid': uid,
			'student': None,
			'status': 'student_not_found',
			'remaining': None,
			'timestamp': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
		}, 30)
		return JsonResponse({'error': 'student not found for tag'}, status=404)

	account, _ = Account.objects.get_or_create(student=student)
	if account.balance < fare:
		RideLog.objects.create(student=student, fare=fare, success=False)
		cache.set('last_scan', {
			'uid': uid,
			'student': student.name,
			'status': 'insufficient_funds',
			'remaining': str(account.balance),
			'timestamp': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
		}, 30)
		return JsonResponse({'status': 'insufficient_funds', 'balance': str(account.balance)}, status=402)

	account.balance -= fare
	account.save()
	RideLog.objects.create(student=student, fare=fare, success=True)

	# store last scan in cache for dashboard polling (short-lived)
	cache.set('last_scan', {
		'uid': uid,
		'student': student.name,
		'status': 'ok',
		'remaining': str(account.balance),
		'timestamp': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
	}, 30)

	return JsonResponse({'status': 'ok', 'remaining': str(account.balance)})


@csrf_exempt
@require_http_methods(["POST"])
def update_cost(request):
	"""Update the cost per ride from the dashboard form. CSRF is exempted so the inline dashboard button works in demo mode."""
	cost = request.POST.get('cost')
	try:
		new_cost = Decimal(str(cost))
	except Exception:
		messages.error(request, 'Invalid cost value')
		return redirect('config:dashboard')

	cfg = SystemConfig.get_solo()
	cfg.cost_per_ride = new_cost
	cfg.save()
	messages.success(request, f'Cost per ride updated to ${new_cost}')
	return redirect('config:dashboard')


@require_http_methods(["GET", "POST"])
def api_cost(request):
	"""GET /api/cost/ - Get the current cost per ride from database.
	POST /api/cost/ - Update cost per ride via AJAX.
	
	GET Response: { 'cost': 25.00 }
	POST Body: { 'cost': 25.00 } (JSON or form-encoded)
	POST Response: { 'status': 'ok', 'cost': 25.00 }
	
	This allows real-time updates across all connected clients without page reload.
	"""
	if request.method == 'GET':
		cfg = SystemConfig.get_solo()
		cost = float(cfg.cost_per_ride)
		min_balance = float(cfg.min_balance)
		return JsonResponse({'cost': cost, 'min_balance': min_balance})

	# POST: Update cost
	try:
		if request.content_type == 'application/json':
			payload = json.loads(request.body.decode('utf-8'))
			cost = payload.get('cost')
		else:
			cost = request.POST.get('cost')
	except Exception:
		return JsonResponse({'status': 'error', 'message': 'Invalid payload'}, status=400)

	if cost is None or cost == '':
		return JsonResponse({'status': 'error', 'message': 'Cost is required'}, status=400)

	# validate and set cost
	try:
		new_cost = Decimal(str(cost))
		if new_cost < 0:
			return JsonResponse({'status': 'error', 'message': 'Cost cannot be negative'}, status=400)
	except Exception:
		return JsonResponse({'status': 'error', 'message': 'Invalid cost value'}, status=400)

	# optional: allow updating min_balance in same request
	min_balance = None
	try:
		if request.content_type == 'application/json':
			payload = json.loads(request.body.decode('utf-8'))
			min_balance = payload.get('min_balance') if 'min_balance' in payload else None
		else:
			min_balance = request.POST.get('min_balance')
	except Exception:
		min_balance = None

	cfg = SystemConfig.get_solo()
	cfg.cost_per_ride = new_cost
	# validate and set min_balance if provided
	if min_balance not in (None, '', []):
		try:
			new_min = Decimal(str(min_balance))
			cfg.min_balance = new_min
		except Exception:
			return JsonResponse({'status': 'error', 'message': 'Invalid min_balance value'}, status=400)

	cfg.save()

	return JsonResponse({'status': 'ok', 'cost': float(new_cost), 'min_balance': float(cfg.min_balance)}, status=200)


@require_http_methods(["GET", "POST"])
def student_register(request):
	if request.method == 'POST':
		name = request.POST.get('name')
		uid = request.POST.get('uid')
		username = request.POST.get('username')
		email = request.POST.get('email')
		balance = request.POST.get('balance')
		grade = request.POST.get('grade')
		roll = request.POST.get('roll')
		parent = request.POST.get('parent')
		if not name or not uid:
			messages.error(request, 'Name and UID required')
			return redirect('config:student_register')

		tag, created = RFIDTag.objects.get_or_create(uid=uid)
		tag.assigned = True
		tag.save()

		# create related User if provided
		user_obj = None
		if username or email:
			uname = username or (email.split('@')[0] if email and '@' in email else None)
			if uname:
				base = uname
				suffix = 0
				while User.objects.filter(username=uname).exists():
					suffix += 1
					uname = f"{base}{suffix}"
				user_obj = User.objects.create(username=uname, email=email or '')
				user_obj.set_unusable_password()
				user_obj.save()

		student = Student.objects.create(name=name, rfid_tag=tag, user=user_obj, grade=grade or None, roll=roll or None, parent_contact=parent or None)
		# initial balance if provided
		try:
			initial_balance = Decimal(str(balance)) if balance not in (None, '', []) else Decimal('0.00')
		except Exception:
			initial_balance = Decimal('0.00')
		Account.objects.create(student=student, balance=initial_balance)
		messages.success(request, f'Student {name} registered')
		return redirect('config:dashboard')

	return render(request, 'config/student_register.html')


@require_http_methods(["GET", "POST"])
def topup(request):
	if request.method == 'POST':
		uid = request.POST.get('uid')
		amount = request.POST.get('amount')
		try:
			amt = Decimal(amount)
		except Exception:
			messages.error(request, 'Invalid amount')
			return redirect('config:topup')

		tag = RFIDTag.objects.filter(uid=uid).first()
		if not tag:
			messages.error(request, 'Tag not found')
			return redirect('config:topup')
		student = Student.objects.filter(rfid_tag=tag).first()
		if not student:
			messages.error(request, 'Student not found for tag')
			return redirect('config:topup')

		account, _ = Account.objects.get_or_create(student=student)
		account.balance += amt
		account.save()
		messages.success(request, f'Added {amt} to {student.name}')
		return redirect('config:dashboard')

	return render(request, 'config/topup.html')


@login_required
def api_topup(request):
	"""AJAX Top Up API: POST /api/topup/ with JSON {uid, amount}"""
	if request.method != 'POST':
		return JsonResponse({'error': 'POST required'}, status=400)
	
	try:
		data = json.loads(request.body)
		uid = data.get('uid', '').strip()
		amount_str = data.get('amount')
		
		if not uid or not amount_str:
			return JsonResponse({'error': 'uid and amount required'}, status=400)
		
		amount = Decimal(str(amount_str))
		if amount <= 0:
			return JsonResponse({'error': 'Amount must be positive'}, status=400)
		
		# Find tag and student
		tag = RFIDTag.objects.filter(uid=uid).first()
		if not tag:
			return JsonResponse({'error': 'Tag not found'}, status=404)
		
		student = Student.objects.filter(rfid_tag=tag).first()
		if not student:
			return JsonResponse({'error': 'Student not found for tag'}, status=404)
		
		# Get or create account and update balance
		account, _ = Account.objects.get_or_create(student=student)
		account.balance += amount
		account.save()
		
		return JsonResponse({
			'success': True,
			'student': student.name,
			'new_balance': float(account.balance),
			'message': f'Added {amount} to {student.name}'
		})
	except Exception as e:
		return JsonResponse({'error': str(e)}, status=500)


def students_page(request):
	students_qs = Student.objects.select_related('rfid_tag').all()
	students_list = []
	for s in students_qs:
		acct = Account.objects.filter(student=s).first()
		students_list.append({
			'id': s.id,
			'name': s.name,
			'rfidUid': s.rfid_tag.uid if s.rfid_tag else '',
			'balance': float(acct.balance) if acct else 0,
			'active': bool(s.rfid_tag and s.rfid_tag.assigned),
			'username': s.user.username if getattr(s, 'user', None) else None,
			'email': s.user.email if getattr(s, 'user', None) else None,
			'grade': s.grade,
			'gender': s.gender,
			'roll': s.roll,
			'parent': s.parent_contact,
		})
	return render(request, 'config/students.html', { 'students_list': students_list })


@require_http_methods(["POST"])
def student_register_ajax(request):
	# Accept form-encoded or JSON
	try:
		if request.content_type == 'application/json':
			payload = json.loads(request.body.decode('utf-8'))
			name = payload.get('name')
			uid = payload.get('uid')
		else:
			name = request.POST.get('name')
			uid = request.POST.get('uid')
	except Exception:
		return JsonResponse({'status': 'error', 'message': 'Invalid payload'}, status=400)

	if not name:
		return JsonResponse({'status': 'error', 'message': 'Name required'}, status=400)

	# optional fields
	username = payload.get('username') if isinstance(payload, dict) else request.POST.get('username')
	email = payload.get('email') if isinstance(payload, dict) else request.POST.get('email')
	balance = payload.get('balance') if isinstance(payload, dict) else request.POST.get('balance')
	grade = payload.get('grade') if isinstance(payload, dict) else request.POST.get('grade')
	gender = payload.get('gender') if isinstance(payload, dict) else request.POST.get('gender')
	roll = payload.get('roll') if isinstance(payload, dict) else request.POST.get('roll')
	# accept new fields if provided
	register_no = payload.get('register_no') if isinstance(payload, dict) else request.POST.get('register_no')
	section = payload.get('section') if isinstance(payload, dict) else request.POST.get('section')
	parent = payload.get('parent') if isinstance(payload, dict) else request.POST.get('parent')
	active = payload.get('active') if isinstance(payload, dict) else request.POST.get('active')

	try:
		initial_balance = Decimal(str(balance)) if balance not in (None, '', []) else Decimal('0.00')
	except Exception:
		initial_balance = Decimal('0.00')

	tag = None
	if uid:
		uid = str(uid).strip()
		tag, created = RFIDTag.objects.get_or_create(uid=uid)
		tag.assigned = bool(active) if active is not None else True
		tag.save()

		# prevent assigning the same RFID tag to multiple students (OneToOne constraint)
		existing_student = Student.objects.filter(rfid_tag=tag).first()
		if existing_student:
			return JsonResponse({'status': 'error', 'message': f'RFID tag {tag.uid} already assigned to {existing_student.name}.'}, status=400)

	user_obj = None
	if username or email:
		uname = username or (email.split('@')[0] if email and '@' in email else None)
		if uname:
			# ensure unique username
			base = uname
			suffix = 0
			while User.objects.filter(username=uname).exists():
				suffix += 1
				uname = f"{base}{suffix}"
			user_obj = User.objects.create(username=uname, email=email or '')
			user_obj.set_unusable_password()
			user_obj.save()

	student = Student.objects.create(name=name, rfid_tag=tag if tag is not None else None, user=user_obj)
	# save metadata fields on student record (grade, roll/register_no, parent_contact/section)
	if grade:
		student.grade = grade
	if gender:
		student.gender = gender
	if register_no:
		student.roll = register_no
	elif roll:
		student.roll = roll
	if section:
		student.parent_contact = section
	elif parent:
		student.parent_contact = parent
	student.save()
	acct = Account.objects.create(student=student, balance=initial_balance)

	# attach extra metadata to response (grade, roll/register_no, section/parent) even if not stored in model
	data = {
		'id': student.id,
		'name': student.name,
		'rfidUid': tag.uid if tag is not None else '',
		'balance': float(acct.balance),
		'active': bool(tag.assigned),
		'username': user_obj.username if user_obj else None,
		'email': user_obj.email if user_obj else None,
		'grade': student.grade,
		'gender': student.gender,
		'register_no': student.roll,
		'section': student.parent_contact,
	}
	return JsonResponse({'status': 'ok', 'student': data}, status=201)


@require_http_methods(["POST"])
def student_update_ajax(request):
	"""Update an existing student's metadata (AJAX)."""
	try:
		payload = json.loads(request.body.decode('utf-8')) if request.content_type == 'application/json' else request.POST
	except Exception:
		return JsonResponse({'status': 'error', 'message': 'Invalid payload'}, status=400)

	sid = payload.get('id') or payload.get('student_id')
	if not sid:
		return JsonResponse({'status': 'error', 'message': 'Student id required'}, status=400)

	student = Student.objects.filter(id=sid).first()
	if not student:
		return JsonResponse({'status': 'error', 'message': 'Student not found'}, status=404)

	# fields to update
	name = payload.get('name') or payload.get('full_name')
	uid = payload.get('uid')
	username = payload.get('username')
	email = payload.get('email')
	balance = payload.get('balance')
	grade = payload.get('grade')
	gender = payload.get('gender')
	roll = payload.get('roll')
	register_no = payload.get('register_no')
	section = payload.get('section')
	parent = payload.get('parent')

	if name:
		student.name = name

	# handle RFID tag assignment
	if uid:
		tag, _ = RFIDTag.objects.get_or_create(uid=uid)
		tag.assigned = True
		tag.save()
		student.rfid_tag = tag

	# handle user creation/update
	user_obj = student.user
	if username or email:
		if user_obj:
			if username:
				user_obj.username = username
			if email:
				user_obj.email = email
			user_obj.save()
		else:
			# create user ensuring unique username
			uname = username or (email.split('@')[0] if email and '@' in email else None)
			if uname:
				base = uname
				suffix = 0
				while User.objects.filter(username=uname).exists():
					suffix += 1
					uname = f"{base}{suffix}"
				user_obj = User.objects.create(username=uname, email=email or '')
				user_obj.set_unusable_password()
				user_obj.save()
				student.user = user_obj

	# update other metadata
	if grade is not None:
		student.grade = grade or None
	if gender is not None:
		student.gender = gender or None
	# map incoming fields to existing model fields
	if register_no is not None:
		student.roll = register_no or None
	elif roll is not None:
		student.roll = roll or None
	if section is not None:
		student.parent_contact = section or None
	elif parent is not None:
		student.parent_contact = parent or None

	student.save()

	# update balance if provided
	if balance is not None:
		try:
			acct = Account.objects.get(student=student)
		except Account.DoesNotExist:
			acct = Account.objects.create(student=student, balance=Decimal('0.00'))
		try:
			acct.balance = Decimal(str(balance))
			acct.save()
		except Exception:
			pass

	acct = Account.objects.filter(student=student).first()

	data = {
		'id': student.id,
		'name': student.name,
		'rfidUid': student.rfid_tag.uid if student.rfid_tag else '',
		'balance': float(acct.balance) if acct else 0,
		'active': bool(student.rfid_tag and student.rfid_tag.assigned),
		'username': student.user.username if getattr(student, 'user', None) else None,
		'email': student.user.email if getattr(student, 'user', None) else None,
		'grade': student.grade,
		'gender': student.gender,
		'register_no': student.roll,
		'section': student.parent_contact,
	}
	return JsonResponse({'status': 'ok', 'student': data})


@require_http_methods(["POST"])
def student_delete_ajax(request):
	try:
		payload = json.loads(request.body.decode('utf-8')) if request.content_type == 'application/json' else request.POST
	except Exception:
		return JsonResponse({'status': 'error', 'message': 'Invalid payload'}, status=400)

	sid = payload.get('id') or payload.get('student_id')
	if not sid:
		return JsonResponse({'status': 'error', 'message': 'Student id required'}, status=400)

	student = Student.objects.filter(id=sid).first()
	if not student:
		return JsonResponse({'status': 'error', 'message': 'Student not found'}, status=404)

	# unassign tag if present
	if student.rfid_tag:
		try:
			tag = student.rfid_tag
			tag.assigned = False
			tag.save()
		except Exception:
			pass

	# delete related account and ride logs
	Account.objects.filter(student=student).delete()
	RideLog.objects.filter(student=student).delete()

	student.delete()
	return JsonResponse({'status': 'ok', 'message': 'Deleted'})


@require_http_methods(["POST"])
def toggle_active(request):
	"""Toggle the active state (RFID tag assigned) for a student's device.
	Expects JSON { "id": <student_id> } or { "uid": <rfid_uid> } or form-encoded POST.
	Returns JSON { 'status': 'ok', 'active': true/false, 'student_name': <name> }.
	"""
	try:
		payload = json.loads(request.body.decode('utf-8')) if request.content_type == 'application/json' else request.POST
	except Exception:
		return JsonResponse({'status': 'error', 'message': 'Invalid payload'}, status=400)

	sid = payload.get('id') or payload.get('student_id')
	uid = payload.get('uid')
	
	if not sid and not uid:
		return JsonResponse({'status': 'error', 'message': 'Student id or UID required'}, status=400)

	# Find student by ID or UID
	if sid:
		student = Student.objects.filter(id=sid).first()
	elif uid:
		tag = RFIDTag.objects.filter(uid=uid).first()
		if not tag:
			return JsonResponse({'status': 'error', 'message': f'RFID tag with UID {uid} not found'}, status=404)
		student = Student.objects.filter(rfid_tag=tag).first()
	
	if not student:
		return JsonResponse({'status': 'error', 'message': 'Student not found'}, status=404)

	if not student.rfid_tag:
		return JsonResponse({'status': 'error', 'message': 'Student has no RFID tag assigned'}, status=400)

	tag = student.rfid_tag
	tag.assigned = not bool(tag.assigned)
	tag.save()

	return JsonResponse({'status': 'ok', 'active': bool(tag.assigned), 'student_name': student.name})


def last_scan(request):
	"""Return the last RFID scan stored in cache (for frontend polling)."""
	data = cache.get('last_scan') or {}
	return JsonResponse({'last_scan': data})


def api_dashboard_data(request):
	"""Return students and recent transactions as JSON for dashboard JS polling."""
	students_qs = Student.objects.select_related('rfid_tag').all()
	students_list = []
	for s in students_qs:
		acct = Account.objects.filter(student=s).first()
		students_list.append({
			'id': s.id,
			'name': s.name,
			'full_name': s.name,
			'rfidUid': s.rfid_tag.uid if s.rfid_tag else '',
			'balance': float(acct.balance) if acct else 0,
			'active': bool(s.rfid_tag and s.rfid_tag.assigned),
			'username': s.user.username if getattr(s, 'user', None) else None,
			'email': s.user.email if getattr(s, 'user', None) else None,
			'grade': s.grade,
			# expose new keys expected by frontend; map to existing fields until DB migration
			'gender': s.gender,
			'register_no': s.roll,
			'section': s.parent_contact,
		})

	recent_qs = RideLog.objects.select_related('student').order_by('-timestamp')[:50]
	recent_list = []
	for r in recent_qs:
		recent_list.append({
			'studentName': r.student.name,
			'amount': float(-r.fare) if r.success else float(-r.fare),
			'description': 'RFID tap',
			'timestamp': timezone.localtime(r.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
			'type': 'payment' if r.success else 'failed'
		})

	return JsonResponse({'students': students_list, 'transactions': recent_list})


@require_http_methods(["GET"])
def api_unregistered_tags(request):
	"""GET /api/unregistered-tags/ - Returns all unregistered tags with timestamps."""
	tags = UnregisteredTag.objects.all().order_by('-timestamp')[:100]
	tags_list = []
	for tag in tags:
		tags_list.append({
			'uid': tag.uid,
			'timestamp': timezone.localtime(tag.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
			'id': tag.id,
		})
	return JsonResponse({'unregistered_tags': tags_list})


@require_http_methods(["GET"])
def api_unregistered_tags_count(request):
	"""GET /api/unregistered-tags/count/ - Returns count of unregistered tags."""
	count = UnregisteredTag.objects.count()
	return JsonResponse({'count': count})


@require_http_methods(["POST"])
def api_unregistered_tags_delete(request):
	"""POST /api/unregistered-tags/delete/ - Delete unregistered tag by ID or clear all."""
	try:
		payload = json.loads(request.body.decode('utf-8')) if request.content_type == 'application/json' else request.POST
	except Exception:
		return JsonResponse({'status': 'error', 'message': 'Invalid payload'}, status=400)

	tag_id = payload.get('id')
	clear_all = payload.get('clear_all', False)

	if clear_all:
		UnregisteredTag.objects.all().delete()
		return JsonResponse({'status': 'ok', 'message': 'All unregistered tags cleared'})
	elif tag_id:
		try:
			tag = UnregisteredTag.objects.get(id=tag_id)
			tag.delete()
			return JsonResponse({'status': 'ok', 'message': 'Tag deleted'})
		except UnregisteredTag.DoesNotExist:
			return JsonResponse({'status': 'error', 'message': 'Tag not found'}, status=404)

	else:
		return JsonResponse({'status': 'error', 'message': 'ID or clear_all required'}, status=400)


@require_http_methods(["POST"])
@login_required
def upload_students_excel(request):
	"""POST /api/upload-students/ - upload .xlsx and create/update student records.

		Expected header columns (case-insensitive):
			name (required), uid (optional)
			username, email, balance, grade, roll, parent, active
	"""
	f = request.FILES.get('file')
	if not f:
		return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)
	try:
		wb = openpyxl.load_workbook(filename=f, read_only=True, data_only=True)
		ws = wb.active
		rows = list(ws.iter_rows(values_only=True))
		if not rows:
			return JsonResponse({'status': 'error', 'message': 'Excel file is empty'}, status=400)

		# normalize headers: lower-case and strip non-alphanumeric characters
		def _norm(s):
			return re.sub(r'[^a-z0-9]', '', str(s).strip().lower()) if s is not None else ''
		# find the header row: some files include a title row before the actual header
		header_row_index = 0
		possible_markers = ('name', 'fullname', 'rfid', 'uid')
		for idx, r in enumerate(rows[:5]):
			# look only at the first few rows for header
			normalized_cells = [ _norm(c) for c in r ]
			if any(any(marker in (cell or '') for cell in normalized_cells) for marker in possible_markers):
				header_row_index = idx
				break
		header = [ _norm(c) for c in rows[header_row_index] ]
		col_index = {name: idx for idx, name in enumerate(header)}

		if header_row_index + 1 >= len(rows):
			return JsonResponse({'status': 'error', 'message': 'Excel file must have a header row and at least one data row'}, status=400)

		created = 0
		updated = 0
		errors = []

		for i, row in enumerate(rows[header_row_index+1:], start=header_row_index+2):
			try:
				def cell(key):
					k = _norm(key)
					idx = col_index.get(k)
					# tolerant matching: if exact normalized key missing, match when key is substring of header or vice-versa
					if idx is None:
						for hname, hidx in col_index.items():
							if not hname:
								continue
							if k == hname or k in hname or hname in k:
								idx = hidx
								break
					return row[idx] if idx is not None and idx < len(row) else None

				name = cell('name') or cell('full name') or cell('fullname')
				uid = cell('uid') or cell('rfid') or cell('rfid uid')
				if not name:
					errors.append({'row': i, 'message': 'Missing name'})
					continue

				username = cell('username')
				email = cell('email')
				balance = cell('balance')
				grade = cell('grade') or cell('class')
				roll = cell('roll')
				# new optional columns we accept but map to existing model fields for now
				register_no = cell('register_no') or cell('register no') or cell('regno')
				section = cell('section')
				gender = cell('gender')
				parent = cell('parent') or cell('parent contact')
				active = cell('active')

				name = str(name).strip()
				tag = None
				if uid is not None and str(uid).strip() != '':
					uid = str(uid).strip()
					tag, _ = RFIDTag.objects.get_or_create(uid=uid)
					if active is None:
						tag.assigned = True
					else:
						a = str(active).strip().lower()
						tag.assigned = not (a in ('0', 'false', 'no', 'n', ''))
					tag.save()
					student = Student.objects.filter(rfid_tag=tag).first()
				else:
					# no UID provided; try to find existing student by name
					student = Student.objects.filter(name__iexact=name).first()
				user_obj = None
				if student:
					student.name = name
					if grade is not None: student.grade = str(grade)
					# prefer register_no if provided, otherwise fall back to roll
					if register_no is not None:
						student.roll = str(register_no)
					elif roll is not None:
						student.roll = str(roll)
					# map section into parent_contact temporarily if provided
					if section is not None:
						student.parent_contact = str(section)
					elif parent is not None:
						student.parent_contact = str(parent)
					if gender is not None:
						student.gender = str(gender)
					student.save()
					updated += 1
				else:
					if username or email:
						uname = username or (str(email).split('@')[0] if email and '@' in str(email) else None)
						if uname:
							base = uname
							suffix = 0
							while User.objects.filter(username=uname).exists():
								suffix += 1
								uname = f"{base}{suffix}"
							user_obj = User.objects.create(username=uname, email=email or '')
							user_obj.set_unusable_password()
							user_obj.save()
					# map register_no -> roll and section -> parent_contact for now
					student = Student.objects.create(
						name=name,
						rfid_tag=tag if tag is not None else None,
						user=user_obj,
						grade=grade or None,
						gender=(str(gender) if gender is not None else None),
						roll=(str(register_no) if register_no is not None else (str(roll) if roll is not None else None)),
						parent_contact=(str(section) if section is not None else (str(parent) if parent is not None else None)),
					)
					created += 1

				try:
					acct = Account.objects.get(student=student)
				except Account.DoesNotExist:
					acct = Account.objects.create(student=student, balance=0)
				# set balance only if provided; accounts default to 0 on creation
				if balance not in (None, ''):
					try:
						acct.balance = Decimal(str(balance))
						acct.save()
					except Exception:
						errors.append({'row': i, 'message': 'Invalid balance value'})

			except Exception as e:
				errors.append({'row': i, 'message': str(e)})

		return JsonResponse({'status': 'ok', 'created': created, 'updated': updated, 'errors': errors})


	except Exception as e:
		return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_http_methods(["GET"])
@login_required
def export_students_excel(request):
	"""GET /api/export-students/ - download a .xlsx backup of all student records."""
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = 'Students'

	headers = [
		'Full name',
		'RFID UID',
		'Username',
		'Email',
		'Balance',
		'Class',
		'Gender',
		'Register No',
		'Section',
		'Active',
	]
	ws.append(headers)

	students_qs = Student.objects.select_related('rfid_tag', 'user').order_by('name')
	for student in students_qs:
		account = Account.objects.filter(student=student).first()
		ws.append([
			student.name,
			student.rfid_tag.uid if student.rfid_tag else '',
			student.user.username if student.user else '',
			student.user.email if student.user else '',
			float(account.balance) if account else 0,
			student.grade or '',
			student.gender or '',
			student.roll or '',
			student.parent_contact or '',
			'Yes' if student.rfid_tag and student.rfid_tag.assigned else 'No',
		])

	for column_cells in ws.columns:
		max_length = max(len(str(cell.value or '')) for cell in column_cells)
		ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 32)

	output = BytesIO()
	wb.save(output)
	output.seek(0)
	filename = f"students-backup-{timezone.localtime().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
	response = HttpResponse(
		output.getvalue(),
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
	)
	response['Content-Disposition'] = f'attachment; filename="{filename}"'
	return response


# ----------------------------------------------------------------
# Bus firmware API endpoints  (WiFi variant -- riverdale-code-wifi)
# ----------------------------------------------------------------

def api_health(request):
	"""GET /api/health/  -- firmware liveness check."""
	return JsonResponse({'status': 'ok'})


def api_bus_sync(request):
	"""GET /api/bus/sync/?bus_id=<id>
	Returns the configured fare and all active students with their server-side balances.
	The bus_id query param is accepted for future per-bus filtering but currently ignored.
	"""
	fare = float(SystemConfig.get_solo().cost_per_ride)
	students_qs = (
		Student.objects
		.select_related('rfid_tag')
		.filter(rfid_tag__isnull=False, rfid_tag__assigned=True)
	)
	students_list = []
	for s in students_qs:
		acct = Account.objects.filter(student=s).first()
		students_list.append({
			'rfid': s.rfid_tag.uid,
			'balance': float(acct.balance) if acct else 0.0,
		})
	return JsonResponse({'fare': fare, 'students': students_list})


@csrf_exempt
def api_bus_transaction(request):
	"""POST /api/bus/transaction/
	Body: {"rfid":"AABBCCDD","fare":25.00,"timestamp":1716000000,"bus_id":"bus1"}
	Logs the ride and deducts the fare from the server-side balance.
	A subsequent wallet-push will override the balance, so this acts as a safe fallback.
	"""
	if request.method != 'POST':
		return JsonResponse({'error': 'POST required'}, status=400)
	try:
		data = json.loads(request.body.decode('utf-8'))
	except Exception:
		return JsonResponse({'error': 'invalid JSON'}, status=400)

	rfid = data.get('rfid')
	fare_raw = data.get('fare')
	ts_unix = data.get('timestamp')
	bus = data.get('bus_id', '')

	if not rfid or fare_raw is None:
		return JsonResponse({'error': 'rfid and fare required'}, status=400)
	try:
		fare = Decimal(str(fare_raw))
	except Exception:
		return JsonResponse({'error': 'invalid fare'}, status=400)

	if ts_unix:
		try:
			ts = datetime.datetime.fromtimestamp(int(ts_unix), tz=datetime.timezone.utc)
		except Exception:
			ts = timezone.now()
	else:
		ts = timezone.now()

	tag = RFIDTag.objects.filter(uid=rfid).first()
	if not tag:
		return JsonResponse({'error': 'tag not found'}, status=404)
	student = Student.objects.filter(rfid_tag=tag).first()
	if not student:
		return JsonResponse({'error': 'student not found'}, status=404)

	account, _ = Account.objects.get_or_create(student=student)
	success = account.balance >= fare
	if success:
		account.balance -= fare
		account.save()

	RideLog.objects.create(student=student, fare=fare, success=success, timestamp=ts, bus_id=bus)

	# Update last_scan cache so the dashboard reflects bus taps in real time
	cache.set('last_scan', {
		'uid': rfid,
		'student': student.name,
		'status': 'ok' if success else 'insufficient_funds',
		'remaining': str(account.balance),
		'timestamp': timezone.localtime(ts).strftime('%Y-%m-%d %H:%M:%S'),
		'bus_id': bus,
	}, 30)

	return JsonResponse({'status': 'ok', 'balance': str(account.balance)}, status=200)


@csrf_exempt
def api_bus_wallets(request):
	"""POST /api/bus/wallets/
	Body: {"bus_id":"bus1","students":[{"rfid":"AABBCCDD","balance":125.00},...] }
	Bulk-updates student balances. The ESP32 is authoritative for balances after each sync.
	"""
	if request.method != 'POST':
		return JsonResponse({'error': 'POST required'}, status=400)
	try:
		data = json.loads(request.body.decode('utf-8'))
	except Exception:
		return JsonResponse({'error': 'invalid JSON'}, status=400)

	students_payload = data.get('students', [])
	if not isinstance(students_payload, list):
		return JsonResponse({'error': 'students must be a list'}, status=400)

	updated = 0
	not_found = 0
	for entry in students_payload:
		rfid = entry.get('rfid')
		balance_raw = entry.get('balance')
		if not rfid or balance_raw is None:
			continue
		try:
			balance = Decimal(str(balance_raw))
		except Exception:
			continue
		tag = RFIDTag.objects.filter(uid=rfid).first()
		if not tag:
			not_found += 1
			continue
		student = Student.objects.filter(rfid_tag=tag).first()
		if not student:
			not_found += 1
			continue
		acct, _ = Account.objects.get_or_create(student=student)
		acct.balance = balance
		acct.save()
		updated += 1

	return JsonResponse({'status': 'ok', 'updated': updated, 'not_found': not_found}, status=200)
