import sys
import re
from decimal import Decimal
import openpyxl


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s).strip().lower()) if s is not None else ''


def cell_for_row(col_index, row, key):
    k = _norm(key)
    idx = col_index.get(k)
    if idx is None:
        for hname, hidx in col_index.items():
            if not hname:
                continue
            if k == hname or k in hname or hname in k:
                idx = hidx
                break
    return row[idx] if idx is not None and idx < len(row) else None


def preview(path, show=10):
    wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print('No rows found')
        return

    # detect header row within first 5 rows
    header_row_index = 0
    possible_markers = ('name', 'fullname', 'rfid', 'uid')
    for idx, r in enumerate(rows[:5]):
        normalized_cells = [_norm(c) for c in r]
        if any(any(marker in (cell or '') for cell in normalized_cells) for marker in possible_markers):
            header_row_index = idx
            break

    header = [_norm(c) for c in rows[header_row_index]]
    col_index = {name: idx for idx, name in enumerate(header)}

    print(f'Detected header row index: {header_row_index}')
    print('Raw header cells:')
    print(rows[header_row_index])
    print('Normalized headers:')
    print(header)

    data_rows = rows[header_row_index+1:header_row_index+1+show]
    print(f'First {len(data_rows)} data rows parsed:')
    for i, row in enumerate(data_rows, start=header_row_index+2):
        name = cell_for_row(col_index, row, 'name') or cell_for_row(col_index, row, 'fullname') or cell_for_row(col_index, row, 'full name')
        uid = cell_for_row(col_index, row, 'uid') or cell_for_row(col_index, row, 'rfid') or cell_for_row(col_index, row, 'rfid uid')
        username = cell_for_row(col_index, row, 'username')
        email = cell_for_row(col_index, row, 'email')
        balance = cell_for_row(col_index, row, 'balance') or cell_for_row(col_index, row, 'initialbalance') or cell_for_row(col_index, row, 'initial balance')
        grade = cell_for_row(col_index, row, 'class') or cell_for_row(col_index, row, 'grade')
        gender = cell_for_row(col_index, row, 'gender')
        register_no = cell_for_row(col_index, row, 'registerno') or cell_for_row(col_index, row, 'register_no') or cell_for_row(col_index, row, 'register no')
        section = cell_for_row(col_index, row, 'section')

        print(f'Row {i}:')
        print('  raw:', row)
        print('  parsed ->', {
            'name': name,
            'uid': uid,
            'username': username,
            'email': email,
            'balance': balance,
            'grade': grade,
            'gender': gender,
            'register_no': register_no,
            'section': section,
        })


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python import_preview.py <path_to_xlsx> [rows]')
        sys.exit(1)
    path = sys.argv[1]
    show = int(sys.argv[2]) if len(sys.argv) > 2 else 10
    preview(path, show=show)
#!/usr/bin/env python3
import os, sys, re, json
from decimal import Decimal

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'myproject.settings')
import django
django.setup()

from config.models import Student, RFIDTag, Account
import openpyxl


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s).strip().lower()) if s is not None else ''


def find_header_row(rows):
    possible_markers = ('name', 'fullname', 'rfid', 'uid')
    for idx, r in enumerate(rows[:6]):
        normalized_cells = [_norm(c) for c in r]
        if any(any(marker in (cell or '') for cell in normalized_cells) for marker in possible_markers):
            return idx
    return 0


def cell_from_row(row, col_index, key):
    k = _norm(key)
    idx = col_index.get(k)
    if idx is None:
        for hname, hidx in col_index.items():
            if not hname:
                continue
            if k == hname or k in hname or hname in k:
                idx = hidx
                break
    return row[idx] if idx is not None and idx < len(row) else None


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: import_preview.py <path-to-xlsx>')
        sys.exit(2)
    path = sys.argv[1]
    if not os.path.exists(path):
        print('File not found:', path)
        sys.exit(2)

    wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print('No rows found')
        sys.exit(1)

    header_row_index = find_header_row(rows)
    header = [_norm(c) for c in rows[header_row_index]]
    col_index = {name: idx for idx, name in enumerate(header)}

    preview = []
    created = 0
    updated = 0
    errors = []

    for i, row in enumerate(rows[header_row_index+1:header_row_index+1+20], start=header_row_index+2):
        try:
            name = cell_from_row(row, col_index, 'name') or cell_from_row(row, col_index, 'fullname') or cell_from_row(row, col_index, 'full name')
            uid = cell_from_row(row, col_index, 'uid') or cell_from_row(row, col_index, 'rfid') or cell_from_row(row, col_index, 'rfid uid')
            username = cell_from_row(row, col_index, 'username')
            email = cell_from_row(row, col_index, 'email')
            balance = cell_from_row(row, col_index, 'initial balance') or cell_from_row(row, col_index, 'balance')
            grade = cell_from_row(row, col_index, 'class') or cell_from_row(row, col_index, 'grade')
            gender = cell_from_row(row, col_index, 'gender')
            register_no = cell_from_row(row, col_index, 'registernumber') or cell_from_row(row, col_index, 'registerno') or cell_from_row(row, col_index, 'registerno') or cell_from_row(row, col_index, 'registerno') or cell_from_row(row, col_index, 'registerno')
            section = cell_from_row(row, col_index, 'section')

            name_s = str(name).strip() if name is not None else ''
            uid_s = str(uid).strip() if uid is not None and str(uid).strip() != '' else None

            # decide whether would create or update
            if uid_s:
                tag = RFIDTag.objects.filter(uid=uid_s).first()
                student = Student.objects.filter(rfid_tag=tag).first() if tag else None
            else:
                student = Student.objects.filter(name__iexact=name_s).first()

            action = 'update' if student else 'create'
            if action == 'create':
                created += 1
            else:
                updated += 1

            preview.append({
                'row': i,
                'name': name_s,
                'uid': uid_s or '',
                'username': username or '',
                'email': email or '',
                'balance': str(balance) if balance is not None else '',
                'class': grade or '',
                'gender': gender or '',
                'register_no': register_no or '',
                'section': section or '',
                'would': action,
            })
        except Exception as e:
            errors.append({'row': i, 'message': str(e)})

    out = {
        'header_row_index': header_row_index,
        'normalized_header': header,
        'preview_rows': preview,
        'created_guess': created,
        'updated_guess': updated,
        'errors': errors,
    }
    print(json.dumps(out, indent=2))
